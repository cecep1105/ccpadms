"""
Export Rekap Absensi (Attendance Recap) ke .xlsx dgn format RAPI (bukan
sekadar dump data spt CSV) -- judul, filter yg diterapkan, header
berwarna, warna hijau/merah utk jam IN/OUT (SAMA konvensi dgn UI web,
lihat halaman Next.js/template Django), freeze panes, kolom lebar
otomatis.

SENGAJA dibuat MODUL TERPISAH (bukan nempel di api_views.py/views.py yg
SUDAH ADA & JALAN) supaya risiko REGRESI ke fitur tampilan yg SUDAH
teruji minimal -- query & agregasi data DI SINI DUPLIKASI SEBAGIAN dari
AttendanceRecapAPIView (BUKAN dipanggil ulang / diimport dari sana),
TRADE-OFF SADAR: sedikit duplikasi kode drpd risiko sentuh logic yg
sudah terbukti benar.
"""
from collections import defaultdict
from datetime import date, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import transaction
from .views import INDONESIAN_DAYS, _is_in_state, _to_local_time

# --- Palet warna, konsisten dgn tema web (primary teal) & konvensi
# warna IN=hijau/OUT=merah yg SUDAH dipakai di tabel web (text-success/
# text-destructive) ---
COLOR_PRIMARY = 'FF1F7A6C'       # teal gelap -- header utama
COLOR_PRIMARY_LIGHT = 'FFE6F2F0'  # teal sangat muda -- sub-header tanggal
COLOR_ZEBRA = 'FFF7F9F9'         # abu-abu SANGAT muda -- selang-seling baris genap
COLOR_IN = 'FF15803D'            # hijau tua -- jam IN
COLOR_OUT = 'FFB91C1C'           # merah tua -- jam OUT
COLOR_WHITE = 'FFFFFFFF'
COLOR_BORDER = 'FFD1D5DB'

THIN_BORDER = Border(*[Side(style='thin', color=COLOR_BORDER) for _ in range(4)])
FONT_NAME = 'Calibri'

RECAP_TYPE_LABEL = {'all': 'All', 'kantin': 'Kantin', 'driver': 'Driver'}


def _build_base_queryset(recap_type, pin, function, pool, device, date_from, date_to):
    """
    Filter TERSENDIRI (BUKAN import dari api_views.py) tapi LOGIC-nya
    SENGAJA disamakan PERSIS dgn AttendanceRecapAPIView.get() supaya
    hasil export & tampilan web KONSISTEN -- HANYA import fungsi
    kode-function dari services.py (itu SUDAH aman & memang dirancang
    dipakai bersama, BEDA dari query-building yg lebih kompleks & rawan
    regresi kalau di-refactor terburu-buru).
    """
    from . import services

    base_qs = transaction.objects.filter(TTime__date__gte=date_from, TTime__date__lte=date_to)
    if recap_type == 'kantin':
        base_qs = base_qs.filter(Function__in=services.kantin_function_codes())
    elif recap_type == 'driver':
        base_qs = base_qs.filter(Function__in=services.driver_function_codes())
    if pin:
        base_qs = base_qs.filter(UserID__PIN__iregex=pin)
    if function:
        base_qs = base_qs.filter(Function=function)
    if device:
        base_qs = base_qs.filter(SN=device)
    elif pool:
        base_qs = base_qs.filter(SN__DeptID=pool)
    return base_qs


def build_attendance_recap_workbook(*, recap_type, pin, function, pool, device, date_from, date_to, filter_summary_extra=None):
    """
    Bangun & kembalikan `openpyxl.Workbook` (BELUM disimpan ke file/
    response) -- BUKAN PAGINATED, SEMUA baris yg cocok filter diekspor
    sekaligus (BEDA dari tampilan web yg dipaginasi) krn tujuan export
    memang utk dibawa/diarsipkan utuh.
    """
    date_columns = []
    d = date_to
    while d >= date_from:
        date_columns.append(d)
        d -= timedelta(days=1)

    base_qs = _build_base_queryset(recap_type, pin, function, pool, device, date_from, date_to)
    pin_list = sorted(set(base_qs.values_list('UserID__PIN', flat=True)))

    matrix = defaultdict(lambda: defaultdict(lambda: {'in': [], 'out': []}))
    names = {}
    if pin_list:
        detail_qs = (
            base_qs.select_related('UserID')
            .order_by('UserID__PIN', 'TTime')
        )
        for trx in detail_qs:
            trx_pin = trx.UserID.PIN
            names[trx_pin] = trx.UserID.EName
            local_time = _to_local_time(trx.TTime)
            if local_time is None:
                continue
            trx_date = local_time.date()
            if _is_in_state(trx.State):
                matrix[trx_pin][trx_date]['in'].append(local_time)
            else:
                matrix[trx_pin][trx_date]['out'].append(local_time)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Rekap Absensi'

    # --- Kolom: No | PIN | Nama | [tiap tanggal: IN | OUT] ---
    n_fixed_cols = 3
    n_date_cols = len(date_columns) * 2
    total_cols = n_fixed_cols + n_date_cols

    # === Baris 1: Judul (merged, seluruh lebar tabel) ===
    title_text = f'Rekap Absensi - {RECAP_TYPE_LABEL.get(recap_type, recap_type)}'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(name=FONT_NAME, size=14, bold=True, color=COLOR_WHITE)
    title_cell.fill = PatternFill('solid', fgColor=COLOR_PRIMARY)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # === Baris 2: Ringkasan periode & filter yang diterapkan ===
    filter_parts = [f'Periode: {date_from.strftime("%d/%m/%Y")} s.d. {date_to.strftime("%d/%m/%Y")}']
    if pin:
        filter_parts.append(f'PIN: {pin}')
    if function:
        filter_parts.append(f'Function: {function}')
    if filter_summary_extra:
        filter_parts.extend(filter_summary_extra)
    filter_parts.append(f'Total employee: {len(pin_list)}')
    subtitle_text = '  |  '.join(filter_parts)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    subtitle_cell = ws.cell(row=2, column=1, value=subtitle_text)
    subtitle_cell.font = Font(name=FONT_NAME, size=9, italic=True, color='FF4B5563')
    subtitle_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 18

    # === Baris 3 (kosong, spacer visual) ===
    ws.row_dimensions[3].height = 6

    # === Baris 4-5: Header tabel (No/PIN/Nama merge vertikal 2 baris, tiap tanggal merge horizontal 2 kolom) ===
    header_row, subheader_row = 4, 5
    for col_idx, label in enumerate(['No', 'PIN', 'Nama'], start=1):
        ws.merge_cells(start_row=header_row, start_column=col_idx, end_row=subheader_row, end_column=col_idx)
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_WHITE)
        cell.fill = PatternFill('solid', fgColor=COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    col = n_fixed_cols + 1
    for d in date_columns:
        day_name = INDONESIAN_DAYS[d.weekday()]
        date_label = f'{day_name}\n{d.strftime("%d/%m/%Y")}'
        ws.merge_cells(start_row=header_row, start_column=col, end_row=header_row, end_column=col + 1)
        date_cell = ws.cell(row=header_row, column=col, value=date_label)
        date_cell.font = Font(name=FONT_NAME, size=9, bold=True, color=COLOR_WHITE)
        date_cell.fill = PatternFill('solid', fgColor=COLOR_PRIMARY)
        date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for sub_idx, sub_label in enumerate(['IN', 'OUT']):
            sub_cell = ws.cell(row=subheader_row, column=col + sub_idx, value=sub_label)
            sub_cell.font = Font(name=FONT_NAME, size=8, bold=True, color=COLOR_PRIMARY.replace('FF', '', 1))
            sub_cell.fill = PatternFill('solid', fgColor=COLOR_PRIMARY_LIGHT)
            sub_cell.alignment = Alignment(horizontal='center', vertical='center')
        col += 2
    ws.row_dimensions[header_row].height = 30
    ws.row_dimensions[subheader_row].height = 16

    # === Baris data ===
    row_idx = subheader_row + 1
    for i, trx_pin in enumerate(pin_list):
        is_even = i % 2 == 1
        row_fill = PatternFill('solid', fgColor=COLOR_ZEBRA) if is_even else None

        no_cell = ws.cell(row=row_idx, column=1, value=i + 1)
        pin_cell = ws.cell(row=row_idx, column=2, value=trx_pin)
        name_cell = ws.cell(row=row_idx, column=3, value=names.get(trx_pin, ''))
        for c in (no_cell, pin_cell, name_cell):
            c.font = Font(name=FONT_NAME, size=9)
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal='center' if c is not name_cell else 'left', vertical='center')
            if row_fill:
                c.fill = row_fill

        col = n_fixed_cols + 1
        for d in date_columns:
            day_data = matrix[trx_pin].get(d, {'in': [], 'out': []})
            in_times = sorted(day_data['in'])
            out_times = sorted(day_data['out'])
            in_text = in_times[0].strftime('%H:%M') if in_times else '-'
            out_text = out_times[-1].strftime('%H:%M') if out_times else '-'
            if len(in_times) > 1:
                in_text += f' ({len(in_times)})'
            if len(out_times) > 1:
                out_text += f' ({len(out_times)})'

            in_cell = ws.cell(row=row_idx, column=col, value=in_text)
            out_cell = ws.cell(row=row_idx, column=col + 1, value=out_text)
            in_cell.font = Font(name=FONT_NAME, size=9, color=COLOR_IN if in_times else 'FF9CA3AF')
            out_cell.font = Font(name=FONT_NAME, size=9, color=COLOR_OUT if out_times else 'FF9CA3AF')
            for c in (in_cell, out_cell):
                c.border = THIN_BORDER
                c.alignment = Alignment(horizontal='center', vertical='center')
                if row_fill:
                    c.fill = row_fill
            col += 2
        row_idx += 1

    # === Lebar kolom ===
    ws.column_dimensions[get_column_letter(1)].width = 5   # No
    ws.column_dimensions[get_column_letter(2)].width = 12  # PIN
    ws.column_dimensions[get_column_letter(3)].width = 26  # Nama
    for c in range(n_fixed_cols + 1, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 9

    # === Freeze panes: baris di atas data (1-5) & 3 kolom label (No/PIN/Nama) TETAP terlihat saat scroll ===
    ws.freeze_panes = ws.cell(row=subheader_row + 1, column=n_fixed_cols + 1)

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
